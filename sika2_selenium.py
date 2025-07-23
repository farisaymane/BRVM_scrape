"""
Module pour extraire les données historiques BRVM depuis Sika Finance avec Selenium
Simule l'interaction utilisateur : saisie des dates et clic sur OK
Version complète avec Volume FCFA et Variations
"""

import pandas as pd
from datetime import datetime
import time
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
import os

class SikaSeleniumExtractor:
    """
    Classe pour extraire les données historiques de Sika Finance avec Selenium
    Extrait: Date, Open, High, Low, Close, Volume Titres, Volume FCFA, Variation %
    """
    
    def __init__(self, headless=True, chromedriver_path=None):
        """
        Initialise l'extracteur Selenium
        
        Args:
            headless: Si True, le navigateur s'exécute en arrière-plan (sans interface)
            chromedriver_path: Chemin vers chromedriver.exe (None pour auto-détection)
        """
        self.headless = headless
        self.chromedriver_path = chromedriver_path
        self.driver = None
        
    def setup_driver(self):
        """Configure et initialise le driver Chrome"""
        print("🚀 Configuration du navigateur Chrome...")
        
        chrome_options = Options()
        
        if self.headless:
            chrome_options.add_argument("--headless")
            print("👻 Mode headless activé (navigateur invisible)")
        else:
            print("👀 Mode visible activé (tu peux voir le navigateur)")
            
        # Options pour stabilité
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--window-size=1920,1080")
        chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
        
        try:
            if self.chromedriver_path:
                service = Service(self.chromedriver_path)
                self.driver = webdriver.Chrome(service=service, options=chrome_options)
            else:
                # Auto-détection du chromedriver
                self.driver = webdriver.Chrome(options=chrome_options)
                
            print("✅ Navigateur Chrome initialisé avec succès")
            
        except WebDriverException as e:
            print(f"❌ Erreur lors de l'initialisation du navigateur: {e}")
            print("💡 Solutions possibles:")
            print("   1. Installer ChromeDriver: pip install chromedriver-autoinstaller")
            print("   2. Ou télécharger manuellement: https://chromedriver.chromium.org/")
            print("   3. Vérifier que Chrome est installé")
            raise
            
    def close_driver(self):
        """Ferme le navigateur"""
        if self.driver:
            self.driver.quit()
            print("🔚 Navigateur fermé")
            
    def extract_data(self, ticker: str, start_date: str, end_date: str, keep_browser_open: bool = False) -> pd.DataFrame:
        """
        Extrait les données historiques en simulant l'interaction utilisateur
        
        Args:
            ticker: Symbole du titre (ex: "BOAB.bj")
            start_date: Date de début (format YYYY-MM-DD)
            end_date: Date de fin (format YYYY-MM-DD)
            
        Returns:
            DataFrame avec les colonnes: Date, Open, High, Low, Close, Volume_Titres, Volume_FCFA, Variation_Pct
        """
        
        print(f"🎯 === EXTRACTION SELENIUM POUR {ticker} ===")
        print(f"📅 Période: {start_date} à {end_date}")
        
        # Validation des dates
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
        except ValueError as e:
            print(f"❌ Format de date invalide: {e}")
            return pd.DataFrame()
            
        # Configuration du navigateur
        if not self.driver:
            self.setup_driver()
            
        try:
            # 1. Naviguer vers la page historique
            url = f"https://www.sikafinance.com/marches/historiques/{ticker}"
            print(f"🌐 Navigation vers: {url}")
            
            self.driver.get(url)
            time.sleep(3)  # Attendre le chargement initial
            
            # 2. Attendre que la page soit complètement chargée
            wait = WebDriverWait(self.driver, 15)
            
            print("⏳ Attente du chargement de la page...")
            
            # 3. Rechercher et remplir le champ date de début
            print("📅 Recherche des champs de date...")
            
            # Différents sélecteurs possibles pour les champs de date
            date_selectors = [
                "input[type='date']",
                "input[name*='date']", 
                "input[id*='date']",
                "input[class*='date']",
                ".date-input",
                "#dateFrom, #dateTo",
                "[placeholder*='date']",
                "input[name='ctl00$ContentPlaceHolder1$txtFrom']",
                "input[name='ctl00$ContentPlaceHolder1$txtTo']"
            ]
            
            start_input = None
            end_input = None
            
            # Essayer de trouver les champs de date
            for selector in date_selectors:
                try:
                    date_inputs = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    if len(date_inputs) >= 2:
                        start_input = date_inputs[0]
                        end_input = date_inputs[1]
                        print(f"✅ Champs de date trouvés avec sélecteur: {selector}")
                        break
                    elif len(date_inputs) == 1:
                        print(f"⚠️ Un seul champ trouvé avec: {selector}")
                except:
                    continue
                    
            # Si pas trouvé, essayer une approche plus générale
            if not start_input or not end_input:
                print("🔍 Recherche alternative des champs...")
                try:
                    # Recherche par texte proche
                    all_inputs = self.driver.find_elements(By.TAG_NAME, "input")
                    print(f"📝 {len(all_inputs)} champs input trouvés au total")
                    
                    for i, input_elem in enumerate(all_inputs):
                        input_type = input_elem.get_attribute("type")
                        input_name = input_elem.get_attribute("name") or ""
                        input_id = input_elem.get_attribute("id") or ""
                        input_class = input_elem.get_attribute("class") or ""
                        
                        print(f"Input {i}: type={input_type}, name={input_name}, id={input_id}")
                        
                        if input_type in ['date', 'text'] and any(keyword in (input_name + input_id + input_class).lower() 
                                                                 for keyword in ['date', 'from', 'to', 'debut', 'fin', 'txtfrom', 'txtto']):
                            if not start_input:
                                start_input = input_elem
                                print(f"📅 Champ de début trouvé: {input_name or input_id}")
                            elif not end_input:
                                end_input = input_elem
                                print(f"📅 Champ de fin trouvé: {input_name or input_id}")
                                break
                                
                except Exception as e:
                    print(f"⚠️ Erreur lors de la recherche alternative: {e}")
            
            if not start_input or not end_input:
                print("❌ Impossible de trouver les champs de date")
                print("🔍 Analyse de la structure de la page...")
                
                # Debug: sauvegarder le HTML pour analyse
                with open(f"debug_{ticker.replace('.', '_')}.html", "w", encoding="utf-8") as f:
                    f.write(self.driver.page_source)
                print(f"💾 HTML sauvegardé pour analyse dans debug_{ticker.replace('.', '_')}.html")
                
                return pd.DataFrame()
            
            # 4. Remplir les champs de date
            print("✏️ Saisie des dates...")
            
            # Convertir les dates au bon format (DD/MM/YYYY pour Sika Finance)
            start_formatted = start_dt.strftime("%d/%m/%Y")
            end_formatted = end_dt.strftime("%d/%m/%Y")
            
            try:
                # Effacer et saisir la date de début
                start_input.clear()
                start_input.send_keys(start_formatted)
                print(f"📅 Date de début saisie: {start_formatted}")
                
                time.sleep(1)
                
                # Effacer et saisir la date de fin
                end_input.clear()
                end_input.send_keys(end_formatted)
                print(f"📅 Date de fin saisie: {end_formatted}")
                
                time.sleep(1)
                
            except Exception as e:
                print(f"❌ Erreur lors de la saisie des dates: {e}")
                return pd.DataFrame()
            
            # 5. Chercher et cliquer sur le bouton OK/Valider
            print("🔍 Recherche du bouton de validation...")
            
            button_selectors = [
                "input[type='submit']",
                "button[type='submit']",
                "input[value='OK']",
                "input[value='Valider']",
                "input[name='ctl00$ContentPlaceHolder1$btnOK']",
                ".btn-submit",
                ".btn-primary",
                "#btnOK",
                "[id*='btnOK']"
            ]
            
            submit_button = None
            
            for selector in button_selectors:
                try:
                    buttons = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    if buttons:
                        submit_button = buttons[0]
                        print(f"✅ Bouton trouvé avec: {selector}")
                        break
                except:
                    continue
            
            # Recherche alternative de bouton
            if not submit_button:
                print("🔍 Recherche alternative du bouton...")
                try:
                    all_buttons = self.driver.find_elements(By.TAG_NAME, "button")
                    all_inputs = self.driver.find_elements(By.CSS_SELECTOR, "input[type='submit'], input[type='button']")
                    
                    all_clickable = all_buttons + all_inputs
                    
                    for btn in all_clickable:
                        btn_text = btn.text.lower()
                        btn_value = (btn.get_attribute("value") or "").lower()
                        btn_name = (btn.get_attribute("name") or "").lower()
                        
                        if any(keyword in btn_text or keyword in btn_value or keyword in btn_name
                               for keyword in ['ok', 'valider', 'rechercher', 'submit', 'chercher', 'btnok']):
                            submit_button = btn
                            print(f"✅ Bouton trouvé: '{btn.text or btn.get_attribute('value')}'")
                            break
                            
                except Exception as e:
                    print(f"⚠️ Erreur recherche bouton: {e}")
            
            if not submit_button:
                print("❌ Bouton de validation non trouvé")
                return pd.DataFrame()
            
            # 6. Cliquer sur le bouton
            print("🖱️ Clic sur le bouton de validation...")
            try:
                self.driver.execute_script("arguments[0].click();", submit_button)
                print("✅ Bouton cliqué avec succès")
            except Exception as e:
                print(f"❌ Erreur lors du clic: {e}")
                try:
                    submit_button.click()
                    print("✅ Clic alternatif réussi")
                except Exception as e2:
                    print(f"❌ Clic alternatif échoué: {e2}")
                    return pd.DataFrame()
            
            # 7. Attendre le rechargement des données
            print("⏳ Attente du rechargement des données...")
            time.sleep(5)  # Attendre que les données se mettent à jour
            
            # 8. Extraire le tableau mis à jour
            print("📊 Extraction du tableau de données...")
            
            # Attendre que le tableau soit présent
            try:
                wait.until(EC.presence_of_element_located((By.TAG_NAME, "table")))
            except TimeoutException:
                print("⚠️ Timeout lors de l'attente du tableau")
            
            # Rechercher le tableau de données
            tables = self.driver.find_elements(By.TAG_NAME, "table")
            
            if not tables:
                print("❌ Aucun tableau trouvé après mise à jour")
                return pd.DataFrame()
            
            print(f"📋 {len(tables)} tableau(x) trouvé(s)")
            
            # Analyser le premier tableau (généralement celui des données)
            table = tables[0]
            
            # Extraire les en-têtes
            headers = []
            try:
                header_row = table.find_element(By.TAG_NAME, "thead").find_element(By.TAG_NAME, "tr")
                header_cells = header_row.find_elements(By.TAG_NAME, "th")
                headers = [cell.text.strip() for cell in header_cells]
                print(f"📋 En-têtes trouvés: {headers}")
            except:
                print("⚠️ En-têtes non trouvés dans thead, essai avec première ligne")
                try:
                    rows = table.find_elements(By.TAG_NAME, "tr")
                    if rows:
                        header_cells = rows[0].find_elements(By.TAG_NAME, "th")
                        if not header_cells:
                            header_cells = rows[0].find_elements(By.TAG_NAME, "td")
                        headers = [cell.text.strip() for cell in header_cells]
                        print(f"📋 En-têtes: {headers}")
                except:
                    print("⚠️ Aucun en-tête trouvé, utilisation d'en-têtes par défaut")
                    headers = ["Date", "Clôture", "Plus bas", "Plus haut", "Ouverture", "Volume Titres", "Volume FCFA", "Variation %"]
            
            # Extraire les lignes de données
            data_rows = []
            try:
                tbody = table.find_element(By.TAG_NAME, "tbody")
                rows = tbody.find_elements(By.TAG_NAME, "tr")
            except:
                rows = table.find_elements(By.TAG_NAME, "tr")
                if headers:  # Ignorer la première ligne si on a trouvé des en-têtes
                    rows = rows[1:]
            
            print(f"📊 {len(rows)} ligne(s) de données trouvées")
            
            # Traiter chaque ligne
            extracted_data = []
            
            for i, row in enumerate(rows):
                try:
                    cells = row.find_elements(By.TAG_NAME, "td")
                    
                    if len(cells) < 7:  # Minimum requis pour les nouvelles colonnes
                        print(f"⚠️ Ligne {i} ignorée: seulement {len(cells)} colonnes")
                        continue
                    
                    cell_values = [cell.text.strip() for cell in cells]
                    print(f"📝 Ligne {i}: {cell_values}")
                    
                    # Essayer de parser les données selon la structure Sika Finance
                    # Colonnes attendues: Date | Clôture | Plus bas | Plus haut | Ouverture | Volume Titres | Volume FCFA | Variation %
                    
                    date_str = cell_values[0]
                    
                    # Parser la date
                    date_obj = None
                    for date_format in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"]:
                        try:
                            date_obj = datetime.strptime(date_str, date_format)
                            break
                        except ValueError:
                            continue
                    
                    if not date_obj:
                        print(f"⚠️ Date non parsable: {date_str}")
                        continue
                    
                    try:
                        # Structure selon Sika Finance: Date, Clôture, Plus bas, Plus haut, Ouverture, Volume Titres, Volume FCFA, Variation %
                        close = self._clean_numeric(cell_values[1])
                        low = self._clean_numeric(cell_values[2])
                        high = self._clean_numeric(cell_values[3])
                        open_price = self._clean_numeric(cell_values[4])
                        volume_titres = self._clean_numeric(cell_values[5])
                        volume_fcfa = self._clean_numeric(cell_values[6]) if len(cell_values) > 6 else "0"
                        variation_pct = self.extract_percentage_with_debug(cell_values, 7)
                        
                        extracted_data.append({
                            'Date': date_obj.strftime('%Y-%m-%d'),
                            'Open': float(open_price),
                            'High': float(high),
                            'Low': float(low),
                            'Close': float(close),
                            'Volume_Titres': int(float(volume_titres)),
                            'Volume_FCFA': int(float(volume_fcfa)),
                            'Variation_Pct': float(variation_pct)
                        })
                        
                    except (ValueError, IndexError) as e:
                        print(f"⚠️ Erreur parsing valeurs ligne {i}: {e}")
                        continue
                        
                except Exception as e:
                    print(f"⚠️ Erreur traitement ligne {i}: {e}")
                    continue
            
            print(f"✅ {len(extracted_data)} lignes extraites avec succès")
            
            if not extracted_data:
                print("❌ Aucune donnée valide extraite")
                return pd.DataFrame()
            
            # Créer le DataFrame
            df = pd.DataFrame(extracted_data)
            df = df.sort_values('Date').reset_index(drop=True)
            df['Date'] = pd.to_datetime(df['Date'])
            
            print(f"📊 Données finales: {len(df)} lignes")
            print(f"📅 Période: {df['Date'].min()} à {df['Date'].max()}")
            print(f"💰 Volume FCFA total: {df['Volume_FCFA'].sum():,} FCFA")
            print(f"📈 Variation min/max: {df['Variation_Pct'].min():.2f}% / {df['Variation_Pct'].max():.2f}%")
            
            return df
            
        except Exception as e:
            print(f"❌ Erreur générale: {e}")
            return pd.DataFrame()
        finally:
            if not keep_browser_open:
                self.close_driver()
    
    def _clean_numeric(self, value_str: str) -> str:
        """Nettoie une valeur numérique"""
        if not value_str:
            return "0"
        
        # Enlever espaces et caractères non numériques (sauf virgules, points et tirets)
        cleaned = re.sub(r'[^\d,.\-]', '', value_str.strip())
        
        # Gérer les séparateurs
        if ',' in cleaned and '.' in cleaned:
            # Format avec milliers et décimales (ex: 1,234.56)
            cleaned = cleaned.replace(',', '')
        elif ',' in cleaned:
            # Vérifier si c'est décimal ou milliers
            parts = cleaned.split(',')
            if len(parts) == 2 and len(parts[1]) <= 3:
                # C'est probablement décimal (ex: 123,45)
                cleaned = cleaned.replace(',', '.')
            else:
                # C'est probablement milliers (ex: 1,234,567)
                cleaned = cleaned.replace(',', '')
        
        return cleaned or "0"
    
    def _clean_percentage(self, value_str: str) -> str:
        """Nettoie une valeur de pourcentage et la divise par 100 si nécessaire"""
        if not value_str:
            return "0"

        # Enlever seulement les espaces et le symbole % si présent
        cleaned = value_str.strip()

        # Enlever le symbole % s'il est présent
        if cleaned.endswith('%'):
            cleaned = cleaned[:-1].strip()

        # Enlever les espaces supplémentaires
        cleaned = cleaned.strip()

        # Gérer le cas où la valeur est vide après nettoyage
        if not cleaned:
            return "0"

        # Remplacer la virgule par un point pour la compatibilité Python
        if ',' in cleaned:
            cleaned = cleaned.replace(',', '.')

        # Valider que c'est un nombre valide et diviser par 100
        try:
            value = float(cleaned)
            # Diviser par 100 car Sika affiche 149% au lieu de 1.49%
            return str(value / 100)
        except ValueError:
            print(f"⚠️ Valeur de pourcentage non valide: '{value_str}' -> '{cleaned}'")
            return "0"
    def extract_percentage_with_debug(self, cell_values, index=7):
        """Version avec debug pour voir les valeurs de pourcentage récupérées"""
        if len(cell_values) > index:
            original_value = cell_values[index]
            cleaned_value = self._clean_percentage(original_value)
            print(f"🔍 Debug variation: '{original_value}' -> '{cleaned_value}'")
            return cleaned_value
        return "0"
    def save_to_excel(self, df: pd.DataFrame, ticker: str, filename: str = None):
        """Sauvegarde les données dans un fichier Excel avec mise en forme complète"""
        if df.empty:
            print("❌ Aucune donnée à sauvegarder")
            return
        
        if not filename:
            filename = f"{ticker.replace('.', '_')}_historique_complet_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        try:
            # Créer un writer Excel avec formatage
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Sauvegarder les données dans la feuille principale
                df.to_excel(writer, sheet_name='Données Historiques', index=False)
                
                # Obtenir le workbook et la worksheet pour le formatage
                workbook = writer.book
                worksheet = writer.sheets['Données Historiques']
                
                # Formatage des colonnes
                from openpyxl.utils import get_column_letter
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.styles.numbers import FORMAT_DATE_YYYYMMDD2, FORMAT_NUMBER_00
                
                # Style pour les en-têtes
                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_alignment = Alignment(horizontal='center', vertical='center')
                
                # Style pour les bordures
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Appliquer le style aux en-têtes
                for col_num in range(1, len(df.columns) + 1):
                    col_letter = get_column_letter(col_num)
                    cell = worksheet[f'{col_letter}1']
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                
                # Formatage des données
                for row_num in range(2, len(df) + 2):
                    for col_num in range(1, len(df.columns) + 1):
                        col_letter = get_column_letter(col_num)
                        cell = worksheet[f'{col_letter}{row_num}']
                        cell.border = thin_border
                        
                        # Format spécifique selon le type de colonne
                        col_name = df.columns[col_num - 1]
                        if col_name == 'Date':
                            cell.number_format = 'DD/MM/YYYY'
                            cell.alignment = Alignment(horizontal='center')
                        elif col_name in ['Open', 'High', 'Low', 'Close']:
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal='right')
                        elif col_name in ['Volume_Titres', 'Volume_FCFA']:
                            cell.number_format = '#,##0'
                            cell.alignment = Alignment(horizontal='right')
                        elif col_name == 'Variation_Pct':
                            # Les variations sont déjà en format décimal (0.0149 pour 1.49%)
                            cell.number_format = '0.00%'
                            cell.alignment = Alignment(horizontal='right')
                            # Colorer en vert si positif, rouge si négatif
                            variation_value = df.iloc[row_num - 2][col_name]
                            if variation_value > 0:
                                cell.font = Font(color='00AA00')  # Vert
                            elif variation_value < 0:
                                cell.font = Font(color='CC0000')  # Rouge
                
                # Ajuster la largeur des colonnes
                column_widths = {
                    'Date': 12,
                    'Open': 10,
                    'High': 10,
                    'Low': 10,
                    'Close': 10,
                    'Volume_Titres': 15,
                    'Volume_FCFA': 18,
                    'Variation_Pct': 12
                }
                
                for col_num, col_name in enumerate(df.columns, 1):
                    col_letter = get_column_letter(col_num)
                    width = column_widths.get(col_name, 12)
                    worksheet.column_dimensions[col_letter].width = width
                
                # Ajouter une feuille de résumé statistique enrichie
                if len(df) > 0:
                    # Calculer les statistiques
                    stats_data = {
                        'Statistique': [
                            'Nombre de jours',
                            'Prix min (Close)',
                            'Prix max (Close)',
                            'Prix moyen (Close)',
                            'Volume Titres total',
                            'Volume Titres moyen',
                            'Volume FCFA total',
                            'Volume FCFA moyen',
                            'Variation % moyenne',
                            'Variation % min',
                            'Variation % max',
                            'Première date',
                            'Dernière date'
                        ],
                        'Valeur': [
                            len(df),
                            f"{df['Close'].min():.2f} FCFA",
                            f"{df['Close'].max():.2f} FCFA",
                            f"{df['Close'].mean():.2f} FCFA",
                            f"{df['Volume_Titres'].sum():,}",
                            f"{df['Volume_Titres'].mean():.0f}",
                            f"{df['Volume_FCFA'].sum():,} FCFA",
                            f"{df['Volume_FCFA'].mean():.0f} FCFA",
                            f"{df['Variation_Pct'].mean():.2f}%",
                            f"{df['Variation_Pct'].min():.2f}%",
                            f"{df['Variation_Pct'].max():.2f}%",
                            df['Date'].min().strftime('%d/%m/%Y'),
                            df['Date'].max().strftime('%d/%m/%Y')
                        ]
                    }
                    
                    stats_df = pd.DataFrame(stats_data)
                    stats_df.to_excel(writer, sheet_name='Résumé Statistique', index=False)
                    
                    # Formater la feuille de résumé
                    stats_worksheet = writer.sheets['Résumé Statistique']
                    
                    # Style des en-têtes du résumé
                    for col_num in range(1, 3):
                        col_letter = get_column_letter(col_num)
                        cell = stats_worksheet[f'{col_letter}1']
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = thin_border
                    
                    # Style des données du résumé
                    for row_num in range(2, len(stats_df) + 2):
                        for col_num in range(1, 3):
                            col_letter = get_column_letter(col_num)
                            cell = stats_worksheet[f'{col_letter}{row_num}']
                            cell.border = thin_border
                            if col_num == 1:  # Colonne Statistique
                                cell.font = Font(bold=True)
                                cell.alignment = Alignment(horizontal='left')
                            else:  # Colonne Valeur
                                cell.alignment = Alignment(horizontal='right')
                    
                    # Ajuster les largeurs des colonnes du résumé
                    # Ajuster les largeurs des colonnes du résumé
                    stats_worksheet.column_dimensions['A'].width = 25
                    stats_worksheet.column_dimensions['B'].width = 20
                
            print(f"✅ Données sauvegardées dans: {filename}")
            print(f"📊 Fichier Excel créé avec {len(df)} lignes de données")
            
        except Exception as e:
            print(f"❌ Erreur lors de la sauvegarde: {e}")
    def extract_data_by_quarters(self, ticker: str, start_date: str, end_date: str) -> pd.DataFrame:
        """
    Extrait les données historiques en découpant la période en tranches de 3 mois
    pour contourner la limitation de Sika Finance
    
    Args:
        ticker: Symbole du titre (ex: "BOAB.bj")
        start_date: Date de début (format YYYY-MM-DD)
        end_date: Date de fin (format YYYY-MM-DD)
        
    Returns:
        DataFrame consolidé avec toutes les données
    """
    
        print(f"🎯 === EXTRACTION PAR PÉRIODES DE 3 MOIS POUR {ticker} ===")
        print(f"📅 Période totale: {start_date} à {end_date}")
    
    # Validation des dates
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
        except ValueError as e:
            print(f"❌ Format de date invalide: {e}")
            return pd.DataFrame()
    
    # Vérifier que la date de fin est après la date de début
        if end_dt <= start_dt:
            print("❌ La date de fin doit être après la date de début")
            return pd.DataFrame()
    
    # Calculer les périodes de 3 mois
        periods = self._calculate_quarterly_periods(start_dt, end_dt)
    
        print(f"📊 {len(periods)} période(s) de 3 mois à extraire")
    
    # DataFrame pour consolider toutes les données
        consolidated_df = pd.DataFrame()
    
    # Configuration du navigateur une seule fois
        if not self.driver:
            self.setup_driver()
    
        try:
            for i, (period_start, period_end) in enumerate(periods, 1):
                print(f"\n🔄 === PÉRIODE {i}/{len(periods)} ===")
                print(f"📅 Du {period_start.strftime('%Y-%m-%d')} au {period_end.strftime('%Y-%m-%d')}")
            
                try:
                # Extraire les données pour cette période en gardant le navigateur ouvert
                    period_df = self.extract_data(
                        ticker, 
                        period_start.strftime('%Y-%m-%d'), 
                        period_end.strftime('%Y-%m-%d'),
                        keep_browser_open=True  # IMPORTANT: garder le navigateur ouvert
                    )
                
                    if not period_df.empty:
                        print(f"✅ Période {i}: {len(period_df)} lignes extraites")
                    
                    # Ajouter au DataFrame consolidé
                        if consolidated_df.empty:
                            consolidated_df = period_df.copy()
                        else:
                        # Concaténer en évitant les doublons sur la date
                            consolidated_df = pd.concat([consolidated_df, period_df], ignore_index=True)
                            consolidated_df = consolidated_df.drop_duplicates(subset=['Date']).reset_index(drop=True)
                    
                    else:
                        print(f"⚠️ Période {i}: Aucune donnée extraite")
                
                except Exception as e:
                    print(f"❌ Erreur période {i}: {e}")
                    continue
            
            # Pause entre les extractions pour éviter de surcharger le serveur
                if i < len(periods):  # Pas de pause après la dernière période
                    print("⏳ Pause de 5 secondes...")
                    time.sleep(5)
        
        # Traitement final du DataFrame consolidé
            if not consolidated_df.empty:
            # Trier par date
                consolidated_df = consolidated_df.sort_values('Date').reset_index(drop=True)
            
            # Supprimer les éventuels doublons finaux
                consolidated_df = consolidated_df.drop_duplicates(subset=['Date']).reset_index(drop=True)
            
                print(f"\n✅ === CONSOLIDATION TERMINÉE ===")
                print(f"📊 Total de lignes consolidées: {len(consolidated_df)}")
                print(f"📅 Période finale: {consolidated_df['Date'].min()} à {consolidated_df['Date'].max()}")
                print(f"💰 Volume FCFA total: {consolidated_df['Volume_FCFA'].sum():,} FCFA")
            
                return consolidated_df
            
            else:
                print("❌ Aucune donnée consolidée")
                return pd.DataFrame()
            
        except Exception as e:
            print(f"❌ Erreur générale extraction par périodes: {e}")
            return pd.DataFrame()
        finally:
            # Fermer le navigateur à la fin de toute l'extraction
            self.close_driver()

    def _calculate_quarterly_periods(self, start_dt: datetime, end_dt: datetime) -> list:
        """
    Calcule les périodes de 3 mois entre deux dates
    
    Args:
        start_dt: Date de début
        end_dt: Date de fin
        
    Returns:
        Liste de tuples (date_début, date_fin) pour chaque période
    """
        from dateutil.relativedelta import relativedelta
    
        periods = []
        current_start = start_dt
    
        while current_start < end_dt:
        # Calculer la fin de la période (3 mois après le début)
            current_end = current_start + relativedelta(months=3)
        
        # S'assurer que la fin ne dépasse pas la date de fin demandée
            if current_end > end_dt:
                current_end = end_dt
        
            periods.append((current_start, current_end))
        
        # Préparer la prochaine période (commencer 1 jour après la fin actuelle)
            current_start = current_end + relativedelta(days=1)
    
        return periods

    def extract_and_save_quarterly(self, ticker: str, start_date: str, end_date: str, filename: str = None):
        """
    Méthode complète : extraction par périodes + sauvegarde automatique
    
    Args:
        ticker: Symbole du titre
        start_date: Date de début (YYYY-MM-DD)
        end_date: Date de fin (YYYY-MM-DD)
        filename: Nom du fichier (optionnel)
    """
    
        print(f"🚀 === EXTRACTION COMPLÈTE PAR PÉRIODES ===")
    
    # Extraire les données
        df = self.extract_data_by_quarters(ticker, start_date, end_date)
    
        if not df.empty:
        # Afficher les statistiques
            print(f"\n📈 === STATISTIQUES FINALES ===")
            print(f"Nombre de jours de trading: {len(df)}")
            print(f"Prix min: {df['Close'].min():.2f} FCFA")
            print(f"Prix max: {df['Close'].max():.2f} FCFA")
            print(f"Prix moyen: {df['Close'].mean():.2f} FCFA")
            print(f"Volume FCFA total: {df['Volume_FCFA'].sum():,} FCFA")
            print(f"Variation % moyenne: {df['Variation_Pct'].mean():.2f}%")
        
        # Sauvegarder avec nom personnalisé
            if not filename:
                filename = f"{ticker.replace('.', '_')}_historique_complet_{start_date}_{end_date}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
            self.save_to_excel(df, ticker, filename)
        
            return df
        else:
            print("❌ Aucune donnée à sauvegarder")
            return pd.DataFrame()

    def extract_multiple_tickers_quarterly(self, tickers: list, start_date: str, end_date: str):
        """
    Extrait plusieurs tickers avec la méthode par périodes
    
    Args:
        tickers: Liste des tickers à extraire
        start_date: Date de début
        end_date: Date de fin
    """
    
        print(f"🎯 === EXTRACTION MULTIPLE PAR PÉRIODES ===")
        print(f"📊 {len(tickers)} ticker(s) à extraire")
        print(f"📅 Période: {start_date} à {end_date}")
    
        results = {}
    
        for i, ticker in enumerate(tickers, 1):
            print(f"\n🔄 === TICKER {i}/{len(tickers)}: {ticker} ===")
        
            try:
                df = self.extract_data_by_quarters(ticker, start_date, end_date)
            
                if not df.empty:
                    results[ticker] = df
                
                # Sauvegarder individuellement
                    filename = f"{ticker.replace('.', '_')}_quarterly_{start_date}_{end_date}.xlsx"
                    self.save_to_excel(df, ticker, filename)
                
                    print(f"✅ {ticker}: {len(df)} lignes - Volume FCFA: {df['Volume_FCFA'].sum():,}")
                
                else:
                    print(f"❌ {ticker}: Aucune donnée extraite")
                
            except Exception as e:
                print(f"❌ Erreur pour {ticker}: {e}")
                continue
        
        # Pause entre les tickers
            if i < len(tickers):
                print("⏳ Pause de 5 secondes avant le prochain ticker...")
                time.sleep(5)
    
    # Créer un fichier consolidé de tous les tickers
        if results:
            print(f"\n💾 === CRÉATION DU FICHIER CONSOLIDÉ ===")
        
            combined_filename = f"BRVM_quarterly_extract_{start_date}_{end_date}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
            try:
                with pd.ExcelWriter(combined_filename, engine='openpyxl') as writer:
                # Feuille de résumé
                    summary_data = []
                    for ticker, df in results.items():
                        summary_data.append({
                            'Ticker': ticker,
                            'Nombre_jours': len(df),
                            'Prix_min': df['Close'].min(),
                            'Prix_max': df['Close'].max(),
                            'Prix_moyen': df['Close'].mean(),
                            'Volume_FCFA_total': df['Volume_FCFA'].sum(),
                            'Variation_moyenne': df['Variation_Pct'].mean()
                        })
                
                    summary_df = pd.DataFrame(summary_data)
                    summary_df.to_excel(writer, sheet_name='Résumé', index=False)
                
                # Feuille pour chaque ticker
                    for ticker, df in results.items():
                        sheet_name = ticker.replace('.', '_')[:31]  # Limite Excel
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
            
                print(f"✅ Fichier consolidé créé: {combined_filename}")
            
            except Exception as e:
                print(f"❌ Erreur création fichier consolidé: {e}")
    
        return results

def main():
    """Fonction principale pour tester l'extracteur"""
    print("🎯 === EXTRACTEUR SIKA FINANCE SELENIUM ===")
    print("📊 Extraction complète avec Volume FCFA et Variations")
    
    # Configuration
    ticker = "SICC.CI"  # Exemple: SICOR
    start_date = "2024-01-01"
    end_date = "2025-06-21"
    
    # Créer l'extracteur (headless=False pour voir le navigateur)
    extractor = SikaSeleniumExtractor(headless=False)
    
    try:
        print(f"🚀 Démarrage de l'extraction pour {ticker}")
        
        # Extraire les données
        df = extractor.extract_data(ticker, start_date, end_date)
        
        if not df.empty:
            print(f"\n📊 === RÉSULTATS DE L'EXTRACTION ===")
            print(f"Nombre de lignes: {len(df)}")
            print(f"Colonnes: {list(df.columns)}")
            print(f"Période: {df['Date'].min()} à {df['Date'].max()}")
            
            # Afficher un aperçu des données
            print(f"\n📋 === APERÇU DES DONNÉES ===")
            print(df.head(10).to_string(index=False))
            
            # Statistiques rapides
            print(f"\n📈 === STATISTIQUES RAPIDES ===")
            print(f"Prix de clôture moyen: {df['Close'].mean():.2f} FCFA")
            print(f"Volume FCFA total: {df['Volume_FCFA'].sum():,} FCFA")
            print(f"Volume FCFA moyen: {df['Volume_FCFA'].mean():.0f} FCFA")
            print(f"Variation % moyenne: {df['Variation_Pct'].mean():.2f}%")
            print(f"Plus forte hausse: {df['Variation_Pct'].max():.2f}%")
            print(f"Plus forte baisse: {df['Variation_Pct'].min():.2f}%")
            
            # Sauvegarder en Excel
            extractor.save_to_excel(df, ticker)
            
            # Optionnel: sauvegarder en CSV simple
            csv_filename = f"{ticker.replace('.', '_')}_data.csv"
            df.to_csv(csv_filename, index=False)
            print(f"💾 Sauvegarde CSV: {csv_filename}")
            
        else:
            print("❌ Aucune donnée extraite")
            
    except Exception as e:
        print(f"❌ Erreur dans main(): {e}")
        
    finally:
        # Toujours fermer le navigateur
        extractor.close_driver()
        print("🏁 Extraction terminée")


def extract_multiple_tickers():
    """Fonction pour extraire plusieurs titres en une fois"""
    print("🎯 === EXTRACTION MULTIPLE ===")
    
    # Liste des tickers à extraire
    tickers = [
        "SICC.CI",    # SICOR
        "BOAB.bj",    # Bank of Africa Bénin
        "CIEC.ci",    # CIE Côte d'Ivoire
        "SPHC.ci",    # SAPH CI
        "SAFC.ci"     # SAFCA CI
    ]
    
    start_date = "2024-01-01"
    end_date = "2025-06-21"
    
    # Créer l'extracteur
    extractor = SikaSeleniumExtractor(headless=True)  # Mode silencieux pour traitement multiple
    
    results = {}
    
    try:
        for ticker in tickers:
            print(f"\n🎯 Extraction de {ticker}...")
            
            try:
                df = extractor.extract_data(ticker, start_date, end_date)
                
                if not df.empty:
                    results[ticker] = df
                    print(f"✅ {ticker}: {len(df)} lignes extraites")
                    
                    # Sauvegarder individuellement
                    extractor.save_to_excel(df, ticker)
                    
                else:
                    print(f"❌ {ticker}: Aucune donnée")
                    
            except Exception as e:
                print(f"❌ Erreur pour {ticker}: {e}")
                
            # Pause entre les extractions pour éviter de surcharger le serveur
            time.sleep(2)
            
        # Résumé final
        print(f"\n📊 === RÉSUMÉ DE L'EXTRACTION MULTIPLE ===")
        for ticker, df in results.items():
            print(f"{ticker}: {len(df)} lignes, Volume FCFA total: {df['Volume_FCFA'].sum():,}")
            
        # Créer un fichier combiné
        if results:
            print(f"\n💾 Création du fichier combiné...")
            combined_filename = f"BRVM_historique_combine_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            with pd.ExcelWriter(combined_filename, engine='openpyxl') as writer:
                for ticker, df in results.items():
                    sheet_name = ticker.replace('.', '_')[:31]  # Limite Excel: 31 caractères
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
            print(f"✅ Fichier combiné créé: {combined_filename}")
            
    except Exception as e:
        print(f"❌ Erreur extraction multiple: {e}")
        
    finally:
        extractor.close_driver()


# === UTILITAIRES BONUS ===

def analyze_ticker_performance(ticker: str, period_days: int = 30):
    """Analyse rapide de performance d'un ticker"""
    from datetime import timedelta
    
    end_date = datetime.now().strftime("%Y-%m-%d")
    start_date = (datetime.now() - timedelta(days=period_days)).strftime("%Y-%m-%d")
    
    extractor = SikaSeleniumExtractor(headless=True)
    
    try:
        df = extractor.extract_data(ticker, start_date, end_date)
        
        if not df.empty:
            first_close = df.iloc[0]['Close']
            last_close = df.iloc[-1]['Close']
            performance = ((last_close - first_close) / first_close) * 100
            
            print(f"\n📈 === ANALYSE {ticker} ({period_days} jours) ===")
            print(f"Prix début période: {first_close:.2f} FCFA")
            print(f"Prix fin période: {last_close:.2f} FCFA")
            print(f"Performance: {performance:.2f}%")
            print(f"Volume FCFA moyen: {df['Volume_FCFA'].mean():.0f} FCFA")
            print(f"Volatilité (écart-type var %): {df['Variation_Pct'].std():.2f}%")
            
            return {
                'ticker': ticker,
                'performance_pct': performance,
                'avg_volume_fcfa': df['Volume_FCFA'].mean(),
                'volatility': df['Variation_Pct'].std(),
                'data_points': len(df)
            }
        else:
            print(f"❌ Pas de données pour {ticker}")
            return None
            
    except Exception as e:
        print(f"❌ Erreur analyse {ticker}: {e}")
        return None
    finally:
        extractor.close_driver()


if __name__ == "__main__":
    # Décommenter la fonction que vous souhaitez utiliser:
    
    # 1. Extraction simple d'un ticker
    main()
    
    # 2. Extraction multiple
    # extract_multiple_tickers()
    
    # 3. Analyse de performance
    # result = analyze_ticker_performance("SICC.CI", 60)
    # if result:
    #     print(f"Résultat analyse: {result}")